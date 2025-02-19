from openpyxl import load_workbook

workbook = load_workbook("tmp/file_example_XLSX_1000.xlsx")
sheet = workbook.active
print(sheet.cell(row=2, column=3).value)

for x in sheet.columns:
    print(x)